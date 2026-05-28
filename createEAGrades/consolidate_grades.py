#!/usr/bin/env python3
import os
import sys
import csv
import re
import argparse
import openpyxl
from openpyxl.utils.cell import range_boundaries

'''

Este script lo deje para facilitar un montón la subida de notas de expoandes. El Excel ya nos ayuda a calcular y consolidar los resultados del 
formato de evaluación. Solamente hay que generar la tabla con los nuevos grupos (para el próximo semestre) con el consolidado de las notas y la relación 
grupo-sección a la que pertenecen (para dar ejemplo pueden revisar el Excel "Evaluación expoandes 202610.xlsx" Hoja "ResultadosProfes" tabla "NotasEA").

La idea es que de bloque neon se descargue desde exportar calificaciones un csv con el código, correo y el grupo al que pertenece en expoandes (este campo es
el más importante). Con eso, el script se hace cargo de hacer match con el excel y crear un CSV para subir las notas a bn! y ya con eso no toca hacerlo manual
jejeje.

'''


def normalize_val(val):
    """Normalize values for robust string comparison."""
    if val is None:
        return ""
    # Convert to string, convert to lowercase, and strip leading/trailing spaces
    s = str(val).strip().lower()
    # Strip trailing .0 if it's a float representation of an integer (e.g. "1.0" -> "1")
    if s.endswith(".0"):
        s = s[:-2]
    return s

def parse_expoandes_field(field_value):
    """
    Extract section and group from ExpoAndes field.
    Matches formats like:
      - 'ExpoaAndes Sección 2 - Grupo 7'
      - 'Sección 5 - Grupo Beta'
    Returns (section, group) tuple or (None, None) if no match.
    """
    if not field_value:
        return None, None
    
    # Regular expression to extract Section and Group
    # Matches 'Sección' or 'Seccion' followed by a number/word, and 'Grupo' followed by any word/number
    pattern = r"secci[oó]n\s+(\w+)\s*-\s*grupo\s+(.+)"
    match = re.search(pattern, field_value, re.IGNORECASE)
    
    if match:
        section = match.group(1).strip()
        group = match.group(2).strip()
        return section, group
    return None, None

def load_excel_grades(excel_path):
    """
    Loads sheet 'ResultadosProfes', finds 'NotasEA' table,
    and returns a mapping of (normalized_section, normalized_group) -> (poster_grade, feria_grade).
    """
    print(f"[*] Opening Excel file: {excel_path}...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"[!] Error reading Excel file: {e}", file=sys.stderr)
        sys.exit(1)
        
    sheet_name = "ResultadosProfes"
    if sheet_name not in wb.sheetnames:
        print(f"[!] Error: Sheet '{sheet_name}' not found in the Excel workbook.", file=sys.stderr)
        sys.exit(1)
        
    sheet = wb[sheet_name]
    table_name = "NotasEA"
    if table_name not in sheet.tables:
        print(f"[!] Error: Excel Table '{table_name}' not found in sheet '{sheet_name}'.", file=sys.stderr)
        sys.exit(1)
        
    table = sheet.tables[table_name]
    print(f"[+] Found table '{table_name}' in range {table.ref}")
    
    # Convert range reference to boundaries
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    
    # Get column headers
    headers = [sheet.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]
    print(f"[+] Table headers found: {headers}")
    
    # Find column indices (0-indexed relative to min_col)
    try:
        sec_idx = headers.index("Sección")
    except ValueError:
        print("[!] Error: 'Sección' column not found in the table headers.", file=sys.stderr)
        sys.exit(1)
        
    # Handle the potential trailing space in 'Grupo ' as identified in inspection
    group_col_name = "Grupo " if "Grupo " in headers else "Grupo"
    try:
        grp_idx = headers.index(group_col_name)
    except ValueError:
        print(f"[!] Error: '{group_col_name}' column not found in the table headers.", file=sys.stderr)
        sys.exit(1)
        
    try:
        poster_idx = headers.index("Poster")
        feria_idx = headers.index("Feria")
    except ValueError as e:
        print(f"[!] Error: Required grade column not found in table headers. {e}", file=sys.stderr)
        sys.exit(1)
        
    grades_map = {}
    print("[*] Parsing table rows...")
    row_count = 0
    for r in range(min_row + 1, max_row + 1):
        sec_val = sheet.cell(row=r, column=min_col + sec_idx).value
        grp_val = sheet.cell(row=r, column=min_col + grp_idx).value
        poster_val = sheet.cell(row=r, column=min_col + poster_idx).value
        feria_val = sheet.cell(row=r, column=min_col + feria_idx).value
        
        if sec_val is None or grp_val is None:
            continue
            
        norm_sec = normalize_val(sec_val)
        norm_grp = normalize_val(grp_val)
        
        grades_map[(norm_sec, norm_grp)] = (poster_val, feria_val)
        row_count += 1
        
    print(f"[+] Loaded {row_count} grade records from Excel table '{table_name}'.")
    return grades_map

def process_csv_and_join(csv_path, output_path, grades_map):
    """
    Reads the input CSV, joins with the Excel grades map, and outputs the result to output_path.
    """
    print(f"[*] Processing CSV file: {csv_path}...")
    if not os.path.exists(csv_path):
        print(f"[!] Error: CSV file '{csv_path}' does not exist.", file=sys.stderr)
        sys.exit(1)
        
    rows_written = 0
    matches_found = 0
    no_matches = []
    
    # Read the CSV
    with open(csv_path, mode='r', encoding='utf-8') as infile:
        # Detect delimiter or use standard csv.DictReader
        reader = csv.DictReader(infile)
        
        # Prepare output headers
        out_headers = ["OrgDefinedId", "Username", "Poster Expoandes Points Grade", "Feria Points Grade", "End-of-Line"]
        
        print(f"[*] Creating output CSV file: {output_path}...")
        with open(output_path, mode='w', encoding='utf-8', newline='') as outfile:
            writer = csv.DictWriter(outfile, fieldnames=out_headers)
            writer.writeheader()
            
            for row in reader:
                org_id = row.get("OrgDefinedId", "").strip()
                email = row.get("Email", "").strip()
                expoandes_val = row.get("ExpoAndes", "").strip()
                
                # Extract Username from Email (part before @)
                username = ""
                if email and "@" in email:
                    username = email.split("@")[0]
                elif email:
                    username = email
                
                poster_grade = ""
                feria_grade = ""
                
                sec, grp = parse_expoandes_field(expoandes_val)
                if sec and grp:
                    norm_sec = normalize_val(sec)
                    norm_grp = normalize_val(grp)
                    
                    key = (norm_sec, norm_grp)
                    if key in grades_map:
                        poster_grade, feria_grade = grades_map[key]
                        matches_found += 1
                    else:
                        no_matches.append({
                            "Name": f"{row.get('First Name', '')} {row.get('Last Name', '')}".strip(),
                            "Email": email,
                            "ExpoAndes": expoandes_val,
                            "Parsed": f"Section: {sec}, Group: {grp}",
                            "Reason": "Section/Group combo not found in Excel"
                        })
                else:
                    if expoandes_val:
                        no_matches.append({
                            "Name": f"{row.get('First Name', '')} {row.get('Last Name', '')}".strip(),
                            "Email": email,
                            "ExpoAndes": expoandes_val,
                            "Parsed": "Failed to parse",
                            "Reason": "Could not parse Section/Group format"
                        })
                    else:
                        # Empty group/section is normal for some rows
                        pass
                
                # Write matching row
                writer.writerow({
                    "OrgDefinedId": org_id,
                    "Username": username,
                    # Estos son los dos campos que usamos (para las notas), se puede mirar el nombre exacto al bajar el CSV de bloque neon.
                    "Poster Expoandes Points Grade": poster_grade if poster_grade is not None else "",
                    "Feria Points Grade": feria_grade if feria_grade is not None else "",
                    "End-of-Line": "#"
                })
                rows_written += 1
                
    print(f"\n[+] Successfully joined and wrote {rows_written} rows to: {output_path}")
    print(f"[+] Total successful grade matches: {matches_found}")
    if no_matches:
        print(f"[!] Warning: {len(no_matches)} rows had an ExpoAndes value but could not be matched:")
        for idx, item in enumerate(no_matches[:10], 1):
            print(f"  {idx}. Student: {item['Name']} ({item['Email']}) | ExpoAndes: '{item['ExpoAndes']}' | Reason: {item['Reason']} ({item['Parsed']})")
        if len(no_matches) > 10:
            print(f"  ... and {len(no_matches) - 10} more unmatched records.")

def main():
    parser = argparse.ArgumentParser(description="Consolidate student grades from an Excel table to a CSV file.")
    parser.add_argument(
        "-e", "--excel", 
        default="Evaluación expoandes 202610.xlsx", 
        help="Path to the Excel file (default: Evaluación expoandes 202610.xlsx)"
    )
    parser.add_argument(
        "-c", "--csv", 
        default="GradesTest.csv", 
        help="Path to the input CSV file (default: GradesTest.csv)"
    )
    parser.add_argument(
        "-o", "--output", 
        default="consolidated_grades_EA.csv", 
        help="Path to the output CSV file (default: consolidated_grades_EA.csv)"
    )

    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"[!] Error: Excel file '{args.excel}' not found.", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(args.csv):
        print(f"[!] Error: CSV file '{args.csv}' not found.", file=sys.stderr)
        sys.exit(1)

    grades_map = load_excel_grades(args.excel)
    process_csv_and_join(args.csv, args.output, grades_map)

if __name__ == "__main__":
    main()
