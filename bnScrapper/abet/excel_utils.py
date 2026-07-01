import re
from typing import Iterable, List, Sequence, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

def _normalize_row(values: Iterable) -> List:
	if values is None:
		raise ValueError("Values cannot be None")
	return list(values)


def _get_sheet(workbook: Workbook, destiny_sheet: Union[str, Worksheet]) -> Worksheet:
	if isinstance(destiny_sheet, Worksheet):
		return destiny_sheet

	if destiny_sheet not in workbook.sheetnames:
		raise ValueError(f"Sheet '{destiny_sheet}' does not exist in workbook")

	return workbook[destiny_sheet]


def _build_unique_table_name(workbook: Workbook, base_name: str) -> str:
	cleaned = re.sub(r"[^A-Za-z0-9_]", "_", (base_name or "Table").strip())
	if not cleaned:
		cleaned = "Table"
	if cleaned[0].isdigit():
		cleaned = f"T_{cleaned}"

	used_names = set()
	for sheet_name in workbook.sheetnames:
		sheet = workbook[sheet_name]
		used_names.update(sheet.tables.keys())

	if cleaned not in used_names:
		return cleaned

	idx = 1
	while f"{cleaned}_{idx}" in used_names:
		idx += 1
	return f"{cleaned}_{idx}"


def _create_table_in_sheet(
	workbook: Workbook,
	sheet: Worksheet,
	headers: Sequence,
	table_name: str = None,
) -> Table:
	normalized_headers = _normalize_row(headers)
	if len(normalized_headers) == 0:
		raise ValueError("Headers must contain at least one column")

	if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
		for col_idx, header in enumerate(normalized_headers, start=1):
			sheet.cell(row=1, column=col_idx, value=header)
	else:
		existing_headers = [sheet.cell(row=1, column=i).value for i in range(1, len(normalized_headers) + 1)]
		if existing_headers != normalized_headers:
			raise ValueError("Sheet already has different headers in first row")

	final_col = len(normalized_headers)
	ref = f"A1:{get_column_letter(final_col)}{max(1, sheet.max_row)}"

	table_display_name = _build_unique_table_name(
		workbook,
		table_name or f"{sheet.title}_table",
	)
	table = Table(displayName=table_display_name, ref=ref)
	style = TableStyleInfo(
		name="TableStyleMedium2",
		showFirstColumn=False,
		showLastColumn=False,
		showRowStripes=True,
		showColumnStripes=False,
	)
	table.tableStyleInfo = style
	sheet.add_table(table)

	for cell in sheet[1]:
		if cell.value is not None:
			cell.font = Font(bold=True)

	return table


def create_workbook_with_sheet(
	sheet_name: str = "Sheet1",
	headers: Sequence = None,
	table_name: str = None,
) -> Workbook:
	"""
	Creates a new workbook with one sheet.

	If headers are provided, writes them in the first row and creates a table.
	Returns the workbook reference.
	"""
	workbook = Workbook()
	sheet = workbook.active
	sheet.title = sheet_name

	if headers:
		_create_table_in_sheet(workbook, sheet, headers, table_name)

	return workbook


def append_array_to_table(
	workbook: Workbook,
	destiny_sheet: Union[str, Worksheet],
	values: Sequence,
) -> bool:
	"""
	Appends one row (or multiple rows) at the end of the table in the given sheet.

	Returns True if the append succeeded, otherwise False.
	"""
	success = False
	try:
		sheet = _get_sheet(workbook, destiny_sheet)
		row_values = _normalize_row(values)

		if not row_values:
			return True

		# Allow passing a single row like ["name", "type"] or many rows like
		# [["name", "type"], ["name2", "type2"]].
		is_single_row = not isinstance(row_values[0], (list, tuple))
		rows_to_append = [row_values] if is_single_row else row_values

		table = next(iter(sheet.tables.values()), None)

		if table:
			min_col, min_row, max_col, max_row = range_boundaries(table.ref)
			expected_columns = max_col - min_col + 1
			for row in rows_to_append:
				if len(row) != expected_columns:
					raise ValueError(
						f"Expected {expected_columns} values for table '{table.displayName}', got {len(row)} in row {row}"
					)

		if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
			raise ValueError("Sheet has no headers/table. Create a table before appending rows")

		for row in rows_to_append:
			sheet.append(list(row))

		append_row_index = sheet.max_row

		if table:
			min_col, min_row, max_col, _ = range_boundaries(table.ref)
			table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{append_row_index}"
		success = True
	except Exception as e:
		print(f"Error appending to table: {e}")
	return success


def add_sheet_with_table(
	workbook: Workbook,
	sheet_name: str,
	headers: Sequence,
	table_name: str = None,
) -> Worksheet:
	"""
	Adds a new sheet to the workbook and creates a table with the given headers.

	Returns the created sheet reference.
	"""
	if sheet_name in workbook.sheetnames:
		raise ValueError(f"Sheet '{sheet_name}' already exists")

	sheet = workbook.create_sheet(title=sheet_name)
	_create_table_in_sheet(workbook, sheet, headers, table_name)
	return sheet

from collections import defaultdict
 
 
def extract_student_groups(xlsx_path: str, activities_config: list) -> dict[str, dict[str, list[str]]]:
    """
    Reads an Excel roster and returns a mapping of group categories to their
    groups and the students belonging to each.
 
    Args:
        xlsx_path: Path to the .xlsx file with student/group data.
        
    Returns:
        A dict shaped as:
        {
            "<groupCategory>": {
                "<groupName>": ["Last, First", ...]
            }
        }
    """
 
    # Collect every unique groupName defined across all activities
    group_categories = {a["groupName"] for a in activities_config if "groupName" in a}
 
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
 
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
 
    # Map column index → category name, only for recognized group categories
    category_col: dict[int, str] = {
        idx: header
        for idx, header in enumerate(headers)
        if header in group_categories
    }
 
    last_name_idx = headers.index("Last Name")
    first_name_idx = headers.index("First Name")
 
    result: dict[str, dict[str, list[str]]] = {cat: defaultdict(list) for cat in group_categories}
 
    for row in ws.iter_rows(min_row=2, values_only=True):
        last = row[last_name_idx]
        first = row[first_name_idx]
        if not last or not first:
            continue
        student_id = f"{last}, {first}"
 
        for col_idx, category in category_col.items():
            group_name = row[col_idx]
            if group_name:
                result[category][group_name].append(student_id)
 
    wb.close()
 
    # Convert inner default dicts to plain dicts for a clean return value
    return {cat: dict(groups) for cat, groups in result.items()}