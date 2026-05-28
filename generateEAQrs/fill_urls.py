import argparse
import urllib.parse
from openpyxl import load_workbook


'''
This script creates the pre filled URLs for the evaluation forms for EA.
Its still on going, we need to make more tests the next semester!
'''
def main():
    parser = argparse.ArgumentParser(description="Llena las URLs pre llenadas en el archivo Excel preservando su formato.")
    parser.add_argument("--excel", required=True, help="Ruta al archivo Excel de entrada.")
    parser.add_argument("--base-url", required=True, help="La URL base del formulario (hasta el signo igual).")
    parser.add_argument("--output", help="Ruta de salida (por defecto, sobreescribe el archivo de entrada).")
    
    args = parser.parse_args()
    input_path = args.excel
    output_path = args.output if args.output else input_path
    base_url = args.base_url

    print(f"Cargando archivo Excel: {input_path}")
    wb = load_workbook(input_path)
    
    if 'Evaluación' not in wb.sheetnames:
        raise ValueError("El archivo Excel no contiene la hoja 'Evaluación'.")
        
    ws = wb['Evaluación']
    
    # 1. Encontrar la fila de encabezados (buscamos la fila que contenga la celda "Nombre")
    header_row = None
    for r in range(1, 10): # Buscar en las primeras 10 filas
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=r, column=c).value).strip() == "Nombre":
                header_row = r
                break
        if header_row:
            break
            
    if not header_row:
        raise ValueError("No se pudo encontrar la fila de encabezados (falta la columna 'Nombre').")
        
    # 2. Encontrar las columnas requeridas en esa fila
    col_stand = None
    col_nombre = None
    col_urls = None
    
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell_value = str(ws.cell(row=header_row, column=col).value).strip() if ws.cell(row=header_row, column=col).value else ""
        if cell_value == "Stand":
            col_stand = col
        elif cell_value == "Nombre":
            col_nombre = col
        elif "URLs pre llenadas" in cell_value: # Búsqueda más flexible por si hay espacios extra
            col_urls = col
            
    if not col_stand or not col_nombre or not col_urls:
        raise ValueError(f"No se encontraron todas las columnas necesarias. Stand={col_stand}, Nombre={col_nombre}, URLs={col_urls}")
        
    # 3. Crear la columna "Nombre Formulario" al final de la tabla
    col_nombre_form = None
    # Verificamos si ya existe
    for col in range(1, max_col + 1):
        if str(ws.cell(row=header_row, column=col).value).strip() == "Nombre Formulario":
            col_nombre_form = col
            break
            
    if not col_nombre_form:
        # Si no existe, la ponemos al final
        col_nombre_form = max_col + 1
        ws.cell(row=header_row, column=col_nombre_form).value = "Nombre Formulario"
        
    print("Procesando filas y generando URLs...")
    count = 0
    
    # 4. Iterar sobre las filas de datos
    for row in range(header_row + 1, ws.max_row + 1):
        stand_val = ws.cell(row=row, column=col_stand).value
        nombre_val = ws.cell(row=row, column=col_nombre).value
        
        # Saltar si no hay datos válidos en Stand o Nombre
        if stand_val is None or nombre_val is None:
            continue
            
        # Limpieza de valores (por si pandas u openpyxl leen enteros como float ej: 1.0)
        if isinstance(stand_val, float) and stand_val.is_integer():
            stand_val = int(stand_val)
            
        # Construir Nombre Formulario: "Stand {Stand}-{Nombre}"
        nombre_form = f"Stand {stand_val} - {nombre_val}"
        
        # Escribir el Nombre Formulario
        ws.cell(row=row, column=col_nombre_form).value = nombre_form
        
        # Construir la URL completa: codificando las comillas dobles y el texto
        # urllib.parse.quote maneja caracteres especiales e inserta %20 para espacios y %22 para comillas
        texto_codificado = urllib.parse.quote(f'"{nombre_form}"')
        url_completa = f"{base_url}{texto_codificado}"
        
        # Escribir la URL
        ws.cell(row=row, column=col_urls).value = url_completa
        count += 1
        
    print(f"Se actualizaron {count} registros.")
    print(f"Guardando archivo Excel en: {output_path}")
    wb.save(output_path)
    print("¡Proceso finalizado con éxito!")

if __name__ == "__main__":
    main()
